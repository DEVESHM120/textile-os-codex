"""One-time script to generate demo/master-fabric-list.xlsx for recruiter testing."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = ["Article", "Content", "Construction", "Cuttable Width", "Finish GSM",
           "Dyeing Method", "Finish", "Country", "K1", "K2", "K3"]

ROWS = [
    ["VDM-SS25-001", "100% Cotton",
     "40s x 40s / 120 x 72 Plain", "58\"", "115", "Continuous Dyeing", "Mercerized Wash", "India",
     "HMM15W", "W11001", "C66001"],
    ["VDM-SS25-002", "100% Cotton",
     "60s x 60s / 140 x 80 Dobby", "60\"", "130", "Package Dyeing", "Soft Touch", "India",
     "HMM20W", "W12001", "C78001"],
    ["VDM-AW25-003", "65% Polyester 35% Cotton",
     "20s x 20s / 96 x 56 Twill", "62\"", "200", "Continuous Dyeing", "Resin Finish", "India",
     "HMM10W", "W09001", "C52001"],
    ["VDM-SS25-004", "97% Cotton 3% Elastane",
     "80s x 80s / 160 x 90 Plain", "58\"", "95", "Package Dyeing", "Cotton Dobby Air", "India",
     "HMM25W", "W15001", "C90001"],
    ["VDM-AW25-005", "100% Cotton",
     "30s x 30s / 108 x 58 Twill", "60\"", "175", "Continuous Dyeing", "Soft Touch", "India",
     "HMM12W", "W10001", "C70001"],
]

def main():
    out_path = Path(__file__).parent.parent / "demo" / "master-fabric-list.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fabric List"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx, row in enumerate(ROWS, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    col_widths = [16, 22, 32, 14, 12, 20, 22, 10, 10, 10, 10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(out_path)
    print(f"Created: {out_path}")

if __name__ == "__main__":
    main()
