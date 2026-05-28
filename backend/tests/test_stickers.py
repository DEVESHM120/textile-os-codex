from pathlib import Path

import openpyxl

from services.feasibility import check_fabric
from services.stickers import generate_bulk_stickers, generate_stickers


PROJECT_DIR = Path(__file__).resolve().parents[2]
SAMPLE = PROJECT_DIR / "data" / "samples" / "sample_cloth_card.txt"


def test_generate_sticker_workbook(tmp_path):
    result = check_fabric(SAMPLE.read_text(encoding="utf-8"), source_filename=SAMPLE.name)
    output = generate_stickers(
        result["card"],
        output_dir=tmp_path,
        workflow_id=123,
        label_copy={"technical_sheet_text": "Demo technical note.", "risk_summary": "Demo risk."},
    )

    path = tmp_path / output["filename"]
    assert path.exists()
    assert output["preview"]["Article"].startswith("HMM15W/")

    wb = openpyxl.load_workbook(path, read_only=True)
    assert wb.sheetnames == ["Sticker", "Technical Sheet", "Fabric Data"]
    wb.close()


def test_generate_bulk_stickers_with_template(tmp_path):
    master = tmp_path / "master.xlsx"
    template = tmp_path / "template.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["K1", "K2", "K3", "K4", "K5", "Content", "Construction", "Weave", "Width", "Finish GSM"])
    ws.append(["A1", "B2", "C66001", "D4", "CT-DA", "100%CO", "40*40-120*72-150-212", "PLAIN", 150, 212])
    wb.save(master)

    twb = openpyxl.Workbook()
    tws = twb.active
    tws["B2"] = "Article"
    tws["B3"] = "Content"
    tws["B4"] = "Construction"
    tws["B5"] = "GSM"
    twb.save(template)

    output = generate_bulk_stickers(master, output_dir=tmp_path, template_path=template)

    assert output["total"] == 1
    assert output["mapping"][0]["field"] == "VARDHMAN_ARTICLE"
    assert output["preview"][0]["Article"] == "A1/B2/C66001/D4/CT-DA"
    assert (tmp_path / output["filename"]).exists()
