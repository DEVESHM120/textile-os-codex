from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.artifacts import safe_artifact_name


K_COLUMNS = ["k1", "k2", "k3", "k4", "k5", "k6", "k7", "k8", "k9"]
PACKAGE_DYEING_PREFIXES = {"S", "C", "W", "F"}
FINISH_LOOKUP = {
    "CT-DA": "Cotton dobby air finish",
    "ST": "Soft touch finish",
    "RF": "Resin finish",
    "MW": "Mercerized wash finish",
}

DEFAULT_BULK_FIELDS = [
    ("Article", "VARDHMAN_ARTICLE"),
    ("Content", "Content"),
    ("Construction", "Construction"),
    ("Cuttable Width", "Width"),
    ("Finish GSM", "Finish GSM"),
    ("Dyeing Method", "DYEING_METHOD"),
    ("Finish", "FINISH_FULL"),
    ("Country", "COUNTRY"),
]

TEMPLATE_KEYWORD_RULES = [
    (["vardhman"], "VARDHMAN_ARTICLE"),
    (["article"], "VARDHMAN_ARTICLE"),
    (["composition"], "Content"),
    (["content"], "Content"),
    (["construction"], "Construction"),
    (["cuttable", "width"], "Width"),
    (["width"], "Width"),
    (["weight"], "Finish GSM"),
    (["gsm"], "Finish GSM"),
    (["dyeing"], "DYEING_METHOD"),
    (["finish"], "FINISH_FULL"),
    (["country"], "COUNTRY"),
]

CONTENT_LOOKUP = {
    "CO": "Cotton",
    "VI": "Viscose",
    "WO": "Wool",
    "SE": "Silk",
    "LI": "Linen",
    "MA": "Modal",
    "MO": "Modal",
    "TE": "Tencel",
    "TEN": "Tencel",
    "LY": "Lyocell",
    "PL": "Polyester",
    "PES": "Polyester",
    "PA": "Nylon",
    "EA": "Elastane",
    "EL": "Elastane",
    "AC": "Acrylic",
    "OCS": "Organic Cotton",
    "RPET": "Recycled Polyester",
    "RCO": "Recycled Cotton",
}


def generate_stickers(
    card: dict[str, Any],
    *,
    output_dir: Path,
    workflow_id: int,
    label_copy: dict[str, Any] | None = None,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = safe_artifact_name(f"workflow_{workflow_id}_labels", ".xlsx")
    out_path = output_dir / filename

    row = fabric_to_label_row(card, label_copy=label_copy)
    wb = openpyxl.Workbook()
    sticker_ws = wb.active
    sticker_ws.title = "Sticker"
    tech_ws = wb.create_sheet("Technical Sheet")
    data_ws = wb.create_sheet("Fabric Data")

    _write_sticker_sheet(sticker_ws, row)
    _write_technical_sheet(tech_ws, card, row)
    _write_data_sheet(data_ws, card)

    wb.save(out_path)
    return {
        "filename": filename,
        "path": str(out_path),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "preview": row,
    }


def fabric_to_label_row(
    card: dict[str, Any],
    *,
    label_copy: dict[str, Any] | None = None,
) -> dict[str, str]:
    article = build_article_number(card)
    k3 = str(card.get("k3", "")).strip()
    dyeing_method = "Package Dyeing" if k3[:1].upper() in PACKAGE_DYEING_PREFIXES else "Continuous Dyeing"
    finish_code = str(card.get("finish_code") or card.get("k5") or "").strip()
    finish_full = FINISH_LOOKUP.get(finish_code.upper(), finish_code)
    label_copy = label_copy or {}
    return {
        "Article": article or str(card.get("card_ref", "")),
        "Card Ref": str(card.get("card_ref", "")),
        "Customer": str(card.get("customer", "")),
        "Buyer": str(card.get("buyer", "")),
        "Content": str(card.get("composition", "")),
        "Construction": str(card.get("construction", "")),
        "Weave": str(card.get("weave", "")),
        "Cuttable Width": _format_value(card.get("finished_width_inches"), " in"),
        "Finish GSM": _format_value(card.get("gsm"), ""),
        "Dyeing Method": dyeing_method,
        "Finish": finish_full,
        "Country": "India",
        "Risk Copy": str(label_copy.get("risk_summary") or ""),
        "Technical Note": str(label_copy.get("technical_sheet_text") or ""),
    }


def build_article_number(card: dict[str, Any]) -> str:
    parts = []
    for key in K_COLUMNS:
        value = str(card.get(key, "")).strip()
        if value and value not in {"0", "0.0", "None"}:
            parts.append(value)
    return "/".join(parts)


def convert_buyer_file(input_path: Path, *, output_dir: Path, sheet_name: str | None = None) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws_in = wb_in[sheet_name] if sheet_name and sheet_name in wb_in.sheetnames else wb_in[wb_in.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    for raw_row in ws_in.iter_rows(values_only=True):
        first_cell = raw_row[0] if raw_row else None
        parsed = _parse_old_navy_style_cell(first_cell)
        if parsed:
            rows.append(parsed)
    wb_in.close()

    out_name = safe_artifact_name("converted_master", ".xlsx")
    out_path = output_dir / out_name
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Converted Master"
    columns = [
        "SRO Number",
        "Roll No",
        "K1",
        "K2",
        "K3",
        "K4",
        "K5",
        "Construction",
        "Weave",
        "Content",
        "Width",
        "Finish GSM",
    ]
    ws_out.append(columns)
    for row in rows:
        ws_out.append([row.get(col, "") for col in columns])
    _style_table(ws_out)
    wb_out.save(out_path)
    return {
        "filename": out_name,
        "path": str(out_path),
        "rows": len(rows),
        "columns": columns,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def generate_bulk_stickers(
    master_path: Path,
    *,
    output_dir: Path,
    template_path: Path | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Generate bulk stickers from master Excel rows and an optional template.

    This follows the original Sticker-Agent model:
    - read master rows from Excel
    - compute K-code/article/dyeing/finish/content fields
    - infer label-to-field mapping from an uploaded template
    - write all stickers into one stacked Excel workbook
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    headers, rows = read_master_rows(master_path, sheet_name=sheet_name)
    if not rows:
        raise ValueError("No data rows found in master Excel.")

    template = inspect_template(template_path) if template_path else default_template_spec()
    filename = safe_artifact_name("bulk_stickers", ".xlsx")
    out_path = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Stickers"
    preview_ws = wb.create_sheet("Preview Data")
    mapping_ws = wb.create_sheet("Template Mapping")

    enriched_rows = [compute_bulk_row_fields(row) for row in rows]
    _write_bulk_sticker_sheet(ws, enriched_rows, template)
    _write_preview_sheet(preview_ws, enriched_rows[:10])
    _write_mapping_sheet(mapping_ws, template, headers)
    wb.save(out_path)

    preview = [
        {label: str(row.get(field, "")) for label, field in template["fields"]}
        for row in enriched_rows[:10]
    ]
    return {
        "filename": filename,
        "path": str(out_path),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "total": len(rows),
        "success": len(enriched_rows),
        "error_count": 0,
        "headers": headers,
        "mapping": template["mapping"],
        "preview": preview,
    }


def read_master_rows(filepath: Path, *, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    ext = filepath.suffix.lower()
    if ext == ".xlsx":
        return _read_master_xlsx(filepath, sheet_name=sheet_name)
    if ext == ".xls":
        return _read_master_xls(filepath)
    raise ValueError("Master file must be .xls or .xlsx.")


def _read_master_xlsx(filepath: Path, *, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_clean_header(value, index) for index, value in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    for row_index, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row_values):
            continue
        row = {headers[index]: _clean_cell(value) for index, value in enumerate(row_values) if index < len(headers)}
        row["__row_num__"] = row_index
        rows.append(row)
    wb.close()
    return headers, rows


def _read_master_xls(filepath: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_index(0)
    headers = [_clean_header(value, index) for index, value in enumerate(ws.row_values(0))]
    rows: list[dict[str, Any]] = []
    for row_index in range(1, ws.nrows):
        values = ws.row_values(row_index)
        if all(value in ("", None, 0.0) for value in values):
            continue
        row = {headers[index]: _clean_cell(value) for index, value in enumerate(values) if index < len(headers)}
        row["__row_num__"] = row_index + 1
        rows.append(row)
    return headers, rows


def inspect_template(template_path: Path | None) -> dict[str, Any]:
    if not template_path or not template_path.exists():
        return default_template_spec()

    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    mapping = []
    fields = []
    for row in ws.iter_rows():
        for index, cell in enumerate(row):
            if cell.value is None:
                continue
            field = _match_template_label(str(cell.value))
            if not field:
                continue
            value_cell = row[index + 1] if index + 1 < len(row) else None
            value_cell_address = value_cell.coordinate if value_cell else ""
            label = str(cell.value).strip()
            mapping.append(
                {
                    "label": label,
                    "field": field,
                    "label_cell": cell.coordinate,
                    "value_cell": value_cell_address,
                    "confidence": "keyword",
                }
            )
            fields.append((label, field))
    wb.close()

    if not fields:
        return default_template_spec()
    return {"fields": fields, "mapping": mapping, "source": template_path.name}


def default_template_spec() -> dict[str, Any]:
    return {
        "fields": DEFAULT_BULK_FIELDS,
        "mapping": [
            {
                "label": label,
                "field": field,
                "label_cell": f"B{index + 2}",
                "value_cell": f"C{index + 2}",
                "confidence": "default",
            }
            for index, (label, field) in enumerate(DEFAULT_BULK_FIELDS)
        ],
        "source": "default",
    }


def compute_bulk_row_fields(row_dict: dict[str, Any]) -> dict[str, Any]:
    row = {_canonical_header(key): value for key, value in row_dict.items()}
    row.update(row_dict)
    enriched = dict(row_dict)
    enriched["VARDHMAN_ARTICLE"] = _join_k_columns(row)
    content = _get_first(row, "content", "Content", "composition")
    enriched["Content"] = _expand_content(str(content or ""))
    construction = str(_get_first(row, "construction", "Construction") or "").strip()
    weave = str(_get_first(row, "weave", "Weave") or "").strip()
    enriched["Construction"] = f"{construction} {weave}".strip() if weave else construction
    k3 = str(_get_first(row, "k3", "K3") or "").strip()
    if k3[:1].upper() in PACKAGE_DYEING_PREFIXES:
        enriched["DYEING_METHOD"] = "Package Dyeing"
        finish_code = str(_get_first(row, "k5", "K5", "finish_code", "Finish Code") or "").strip()
        enriched["FINISH_FULL"] = FINISH_LOOKUP.get(finish_code.upper(), finish_code)
    else:
        enriched["DYEING_METHOD"] = "Continuous Dyeing"
        enriched["FINISH_FULL"] = ""
    enriched["COUNTRY"] = "India"
    enriched["Width"] = _compute_cuttable_width(row)
    enriched["Finish GSM"] = _get_first(row, "finish_gsm", "Finish GSM", "gsm", "GSM") or ""
    return enriched


def _write_bulk_sticker_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[dict[str, Any]],
    template: dict[str, Any],
) -> None:
    ws["A1"] = "Bulk Sticker Output"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E5F")
    ws.merge_cells("A1:C1")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 72
    border = _thin_border()
    label_font = Font(bold=True)
    start_row = 3
    gap = 1
    block_height = len(template["fields"])
    for sticker_index, row in enumerate(rows, start=1):
        top = start_row + (sticker_index - 1) * (block_height + gap + 1)
        title_cell = ws.cell(row=top, column=2, value=f"Sticker {sticker_index}")
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="36594C")
        ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=3)
        for offset, (label, field) in enumerate(template["fields"], start=1):
            out_row = top + offset
            label_cell = ws.cell(row=out_row, column=2, value=label)
            value_cell = ws.cell(row=out_row, column=3, value=_stringify_excel_value(row.get(field, "")))
            label_cell.font = label_font
            for cell in (label_cell, value_cell):
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_preview_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["No rows"])
        return
    fields = ["VARDHMAN_ARTICLE", "Content", "Construction", "Width", "Finish GSM", "DYEING_METHOD", "FINISH_FULL"]
    ws.append(fields)
    for row in rows:
        ws.append([_stringify_excel_value(row.get(field, "")) for field in fields])
    _style_table(ws)
    for index in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 24 if index != 3 else 56


def _write_mapping_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    template: dict[str, Any],
    headers: list[str],
) -> None:
    ws.append(["Template Label", "Resolved Field", "Label Cell", "Value Cell", "Confidence"])
    for item in template["mapping"]:
        ws.append([
            item["label"],
            item["field"],
            item.get("label_cell", ""),
            item.get("value_cell", ""),
            item.get("confidence", ""),
        ])
    ws.append([])
    ws.append(["Master Columns"])
    for header in headers:
        ws.append([header])
    _style_table(ws)
    for index in range(1, 6):
        ws.column_dimensions[get_column_letter(index)].width = 24


def _clean_header(value: Any, index: int) -> str:
    if value in (None, ""):
        return f"COL_{index + 1}"
    return str(value).strip()


def _clean_cell(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _canonical_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def _match_template_label(label_text: str) -> str | None:
    lower = label_text.lower().strip()
    for keywords, field in TEMPLATE_KEYWORD_RULES:
        if all(keyword in lower for keyword in keywords):
            return field
    return None


def _join_k_columns(row: dict[str, Any]) -> str:
    parts = []
    for index in range(1, 10):
        value = _get_first(row, f"k{index}", f"K{index}")
        value = str(value or "").strip()
        if value and value not in {"0", "0.0", "None"}:
            parts.append(value)
    return "/".join(parts)


def _get_first(row: dict[str, Any], *keys: str) -> Any:
    canonical = {_canonical_header(key): value for key, value in row.items()}
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
        normalized = _canonical_header(key)
        if normalized in canonical and canonical[normalized] not in (None, ""):
            return canonical[normalized]
    return ""


def _expand_content(content: str) -> str:
    if not content.strip():
        return content
    tokens = re.findall(r"(\d+\.?\d*)%(\S+)", content)
    if not tokens:
        return content
    parts = []
    for percent, code in tokens:
        code_clean = code.strip().upper()
        full = CONTENT_LOOKUP.get(code_clean, code)
        try:
            percent_value = float(percent)
            percent_text = str(int(percent_value)) if percent_value.is_integer() else percent
        except ValueError:
            percent_text = percent
        parts.append(f"{percent_text}% {full}")
    return " ".join(parts)


def _compute_cuttable_width(row: dict[str, Any]) -> str:
    raw_width = _get_first(row, "width", "Width", "cuttable_width", "Cuttable Width")
    width_cm = _as_float(raw_width)
    if width_cm is None:
        width_cm = _extract_width_from_construction(
            str(_get_first(row, "construction", "Construction") or ""),
            _get_first(row, "finish_gsm", "Finish GSM", "gsm", "GSM"),
        )
    if width_cm is None:
        return ""
    return str(round((width_cm / 2.54) - 2, 2))


def _extract_width_from_construction(construction: str, gsm_value: Any) -> float | None:
    gsm = _as_float(gsm_value)
    if not construction or gsm is None:
        return None
    candidates = [int(match) for match in re.findall(r"\b(\d+)\b", construction)]
    if not candidates:
        return None
    best_index = None
    best_diff = float("inf")
    for index, number in enumerate(candidates):
        diff = abs(number - gsm)
        if (diff / max(gsm, 1) <= 0.05 or diff <= 2) and diff < best_diff:
            best_index = index
            best_diff = diff
    if best_index is None or best_index == 0:
        return None
    width_candidates = [number for number in candidates[:best_index] if 85 <= number <= 250]
    return float(width_candidates[-1]) if width_candidates else None


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        match = re.search(r"\d+(?:\.\d+)?", str(value))
        return float(match.group(0)) if match else None


def _stringify_excel_value(value: Any) -> str:
    if value in (None, "None"):
        return ""
    return str(value)


def _thin_border() -> Border:
    side = Side(style="thin", color="A6A6A6")
    return Border(left=side, right=side, top=side, bottom=side)


def _parse_old_navy_style_cell(value: Any) -> dict[str, Any] | None:
    if value in (None, ""):
        return None
    parts = [part.strip() for part in str(value).split(",")]
    if len(parts) < 3:
        return None
    k_chain = [part.strip() for part in (parts[3] if len(parts) > 3 else "").split("/") if part.strip()]
    return {
        "SRO Number": parts[0] if len(parts) > 0 else "",
        "Roll No": parts[1] if len(parts) > 1 else "",
        "K1": parts[2] if len(parts) > 2 else "",
        "K2": k_chain[0] if len(k_chain) > 0 else "",
        "K3": k_chain[1] if len(k_chain) > 1 else "",
        "K4": k_chain[2] if len(k_chain) > 2 else "",
        "K5": (k_chain[3] if len(k_chain) > 3 else "") or (parts[10] if len(parts) > 10 else ""),
        "Construction": parts[5] if len(parts) > 5 else "",
        "Content": parts[6] if len(parts) > 6 else "",
        "Width": _extract_number(parts[7] if len(parts) > 7 else ""),
        "Finish GSM": _extract_number(parts[8] if len(parts) > 8 else ""),
        "Weave": parts[9] if len(parts) > 9 else "",
    }


def _write_sticker_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, row: dict[str, str]) -> None:
    ws["A1"] = "Production Sticker"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E5F")
    ws.merge_cells("A1:B1")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 62
    start = 3
    for offset, (label, value) in enumerate(row.items()):
        ws.cell(start + offset, 1, label)
        ws.cell(start + offset, 2, value)
    _style_table(ws, max_row=start + len(row) - 1, max_col=2)


def _write_technical_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    card: dict[str, Any],
    row: dict[str, str],
) -> None:
    ws["A1"] = "Technical Communication Sheet"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="36594C")
    ws.merge_cells("A1:C1")
    ws.append([])
    ws.append(["Field", "Value", "Production Note"])
    fields = [
        ("Gate Data Source", card.get("source_filename", ""), "Generated from feasibility workflow data."),
        ("Article", row.get("Article", ""), "Used for sticker and swatch tracking."),
        ("Construction", row.get("Construction", ""), "Confirm before sampling release."),
        ("GSM", row.get("Finish GSM", ""), "Deterministic feasibility check remains source of truth."),
        ("Width", row.get("Cuttable Width", ""), "Review against buyer requirement."),
        ("Content", row.get("Content", ""), "Must match compliance label."),
        ("AI Technical Note", row.get("Technical Note", ""), "Draft copy for merchandiser/technical communication."),
    ]
    for field in fields:
        ws.append(list(field))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 54
    _style_table(ws, max_row=ws.max_row, max_col=3)


def _write_data_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, card: dict[str, Any]) -> None:
    ws.append(["Field", "Value"])
    for key in sorted(card.keys()):
        ws.append([key, card.get(key)])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    _style_table(ws)


def _style_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    max_row: int | None = None,
    max_col: int | None = None,
) -> None:
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    border = Border(
        left=Side(style="thin", color="A6A6A6"),
        right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"),
        bottom=Side(style="thin", color="A6A6A6"),
    )
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row in {1, 3}:
                cell.font = Font(bold=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _extract_number(raw: Any) -> str:
    match = re.search(r"\d+(?:\.\d+)?", str(raw))
    return match.group(0) if match else ""


def _format_value(raw: Any, suffix: str) -> str:
    if raw in (None, ""):
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return f"{int(raw)}{suffix}"
    return f"{raw}{suffix}"
